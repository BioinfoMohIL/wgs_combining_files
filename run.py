import os

import re

from datetime import datetime

import logging

import warnings

import numpy as np

import pandas as pd

from openpyxl import Workbook, load_workbook

from openpyxl.styles import PatternFill

 

'''

Combine the sending file (WGS_XXXX) with 'star' files :

    [0] Setup:

        -> If .csv, convert to xl (we prefer to work with xl files).

    [1] Fetch 'Samples codes' and 'Production date' columns.

    [2] Concate Run files (Star_0XX) with the sending file (WGS_XXX)

        ->  The 'Samples codes' of the sending files has to be = to

            the 'Source Barcode' of the run file

    [3] Create Tables 96w like, and highlight to sample we need to run from each

        table (in target position)

    [4] Create inputs for running easily Myra program

'''

 

warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')

 

RUN_FILE_PATTERN =  'star'

 

def create_dir(d):

    if not os.path.exists(d):

 

        try:

 

            os.makedirs(d)

 

            logging.info(f"[create_dir] Successfully created directory: {results_dir}\n")

 

        except Exception as e:

 

            logging.error(f"[create_dir] Error occurred while creating directory {d}: {e}\n")

 

            exit()

 

def convert_csv_to_xl(data_dir, file):

    csv_file = os.path.join(data_dir, file)

 

    df = pd.read_csv(csv_file)  

 

    excel_file = os.path.join(data_dir, file.replace('.csv', '.xlsx'))

 

       

 

    df.to_excel(excel_file, index=False)

          

 

def return_extention(f):

    _, file_extension = os.path.splitext(f)

 

    return file_extension

 

def read_table_file(file_name, ext):

    if ext == '.csv':

        df = pd.read_csv(file_name)

    elif ext == '.xlsx':

        df = pd.read_excel(file_name)

    else:

        exit("Unsupported file format")

 

    return df

 

def get_run_files(pattern, extension):

    return [

    filename for filename in os.listdir(data_dir)

    if filename.lower().startswith(RUN_FILE_PATTERN) and filename.lower().endswith(extension)

    ]

 

def fetch_filename(data_folder, file_name):

    file_path = os.path.join(data_folder, file_name)

    file_name = os.path.basename(file_path)

 

    return file_name

 

 

def fetch_columns(df, required_columns):

 

    """

    Fetches columns from the DataFrame by matching them against the list of required columns.

    Handles cases where there are extra spaces, capitalization differences, etc.

 

    Args:

    df: The pandas DataFrame.

    required_columns: A list of required column names to fetch.

 

    Returns:

    A DataFrame with only the required columns, if they are found.

    """

    # Standardize column names in the DataFrame

    df.columns = df.columns.str.strip().str.lower()

 

    # Standardize the required columns list (strip spaces and convert to lowercase)

    required_columns_standardized = [col.strip().lower() for col in required_columns]

 

    # Map the required columns to the actual columns in the DataFrame

    available_columns = []

 

    for col in required_columns_standardized:

        for df_col in df.columns:

            if col in df_col:  # Check if the required column is a substring of the actual column

                available_columns.append(df_col)

                break

 

    if not available_columns:

        raise ValueError("None of the required columns were found.")

 

    return available_columns

 

def replace_hyphen_by_dot(df, col_name):

    pd.Series(df[col_name]).str.replace('-', '.', regex=False)

    return df

 

 

def sort_files(d):

    files = os.listdir(d)

    pattern = r'[Ss]tar_(\d+)_(.+)'  

 

    file_data = []

 

    for file in files:

        match = re.match(pattern, file)

        if match:

            numeric_part = int(match.group(1))

            file_data.append((numeric_part, file))  

 

    file_data.sort(key=lambda x: x[0])  

 

    return [file for _, file in file_data]

 

 

def sort_df(df, col_name):

    df['numeric_part'] = df[col_name].str.extract(r'(\d+)').astype(int)

 

    return df.sort_values(by='numeric_part', ascending=True).drop(columns='numeric_part')

 

 

def create_plate_df():

    df = pd.DataFrame(np.empty((8, 12), dtype=object))

    df.columns = [str(i + 1) for i in range(12)]

    df.index = [chr(65 + i) for i in range(8)]

 

    return df

 

 

# Try to fix the issue of different lower/upper case col header

def manage_columns_typo(df):

    return df.columns.str.strip().str.lower()

 

def create_table_df():

    index = [f"{chr(col)}{row}" for row in range(1, 13) for col in range(ord('A'), ord('H') + 1)]

 

    df = pd.DataFrame(index=index)

    df[SOURCE_NAME_COL] = ''

 

    return df

 

def color_header(ws, color=None):

    for cell in ws[1]:  

        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type='solid')

 

    for cell in ws['A']:

        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type='solid')

 

def cells_to_highlight(df, star_run):

    cells = []

    cells_positions = {}

 

    current_df = df[df[STAR_RUN_COL] == star_run]

 

 

 

    for _, row in current_df.iterrows():

 

        cells.append(row[SAMPLES_CODES_COL])

 

        cells_positions[row[TARGET_POSITION_COL]] = row[SAMPLES_CODES_COL]

 

 

 

    return cells, cells_positions

 

 

 

def populate_table(d, table, col):

 

    for key, value in d.items():

 

        table.loc[key, col] = value

 

 

 

def create_new_tables_sheet(writer, tables, sheet_name):

 

 

 

    rows = []

 

    for key, df in tables.items():

 

        rows.append([key,'',''])

 

        rows.append(['','',SOURCE_NAME_COL])

 

        for index, row in df.iterrows():

 

            rows.append(['', index, row.values[0]])

 

        rows.append(['','',''])

 

    result_df = pd.DataFrame(rows, columns=[None, None, None])

 

    result_df.to_excel(writer, sheet_name=sheet_name, index=False)

 

           

 

def get_position(df, star_run):

 

    cells_positions = []

 

   

 

    current_df = df[df[STAR_RUN_COL] == star_run]

 

 

 

    for _, row in current_df.iterrows():

 

        cells_positions.append(row[TARGET_POSITION_COL])

 

 

 

    return cells_positions

 

 

 

def highlighting_cells(ws, cells):

 

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')        

 

    for row in ws.iter_rows():

 

        for cell in row:

 

            if cell.value in cells:

 

                cell.fill = yellow_fill        

 

def fetch_and_concatenate(df, data_dir, data):

    file_path = os.path.join(data_dir, data)

    file_name = os.path.basename(file_path)

    run_name = file_name.split('_')[1] 

 

    ext = return_extention(file_path)

 

    df_star = read_table_file(file_path, ext)

 

    # Ensure that columns are properly formatted (strip whitespace, lowercase)

    df_star.columns = manage_columns_typo(df_star)

 

    # We need to replace '.' by '-' in the file to have matchs with Star files codes

    # The star files have codes with '-'

    df_star = replace_hyphen_by_dot(df_star, SOURCE_BARCODE_COL)

    

    # Add the Star run

    df_star[STAR_RUN_COL] = run_name

 

    if SOURCE_BARCODE_COL in df_star.columns and SAMPLES_CODES_COL in df.columns:

        for i, sample_code in enumerate(df[SAMPLES_CODES_COL]):

            matched_rows = df_star[df_star[SOURCE_BARCODE_COL].str.lower() == sample_code.lower()]

 

            # If there's a match, append the row to the corresponding row in df_result

            if not matched_rows.empty:

                for column in matched_rows.columns:

                    # Add the matched row to df_result (or merge the columns if already exists)

                    df.loc[i, column] = matched_rows.iloc[0][column]

     

    return df[[SAMPLES_CODES_COL, SOURCE_BARCODE_COL, TARGET_POSITION_COL, STAR_RUN_COL]]

 

 

 

def create_tables(data_folder, df_sorted, output_file):

    tables_by_run = {}              # tables to input Myra

    tables_by_run_cleaned = {}      # tables with dropna

    tables_cleaned_for_plates = {}  # tables cleaned to Myra plates view (96w)

 

    with pd.ExcelWriter(output_file,  engine='openpyxl', mode='a') as writer:

 

        files = sort_files(data_folder)

        for file_name in files:

            cells = []

 

            run_name = file_name.split('_')[1]

 

        

            # Create df like 96w plate A -> H (rows), 1 -> 12 (cols)

 

            df_plate = create_plate_df()

 

            df_table = create_table_df()

 

           

 

            file_name = os.path.join(data_folder, file_name)

 

           

 

            ext = return_extention(file_name)

 

 

 

            df = read_table_file(file_name, ext)

 

            df.columns = manage_columns_typo(df)

 

 

 

            df = replace_hyphen_by_dot(df, SOURCE_BARCODE_COL)

 

            df = df.dropna()  

 

 

 

            for _, row in df.iterrows():

 

                target_position = str(row[TARGET_POSITION_COL])

 

                sample_code = row[SOURCE_BARCODE_COL]

 

               

 

                if target_position:

 

                    plate_row = target_position[0]

 

                    plate_col = target_position[1:]

 

                   

 

                    df_plate.loc[plate_row, plate_col] = sample_code

 

 

 

            cells, positions = cells_to_highlight(df_sorted, run_name)

 

           

 

            df_plate.to_excel(writer, sheet_name=run_name)

 

           

 

            populate_table(positions, df_table, SOURCE_NAME_COL)

 

 

 

            # Color cells

 

            ws = writer.sheets[run_name]

 

            color_header(ws,'ADD8E6')

 

            highlighting_cells(ws, cells)

 

 

 

            # Create another tables for Myra input

 

            tables_by_run[run_name] = df_table

 

            tables_by_run_cleaned[run_name] = df_table.replace([None, ''], pd.NA).dropna()

 

 

 

            tables_cleaned_for_plates[run_name] = cleaning_run_empties(df_table, SOURCE_NAME_COL)

 

   

 

    create_tables_myra_input(tables_by_run, output_file, 'tables')

 

 

 

    create_tables_myra_input(tables_by_run_cleaned, output_file, 'tables_cleaned')

 

 

 

    create_plates_myra(tables_cleaned_for_plates, output_file)

 

           

 

######################################################

 

##### MYRA ###########################################

 

# 1. 'tables' sheet

 

# We want a tables sheet containing all tables for Myra input

 

# The Myra need position (A1, B1...), and the sample next to each position

 

# If we dont have sample for a position, the cell will be empty

 

def create_tables_myra_input(tables, outp, sheetname):    

 

    try:

 

        with pd.ExcelWriter(outp,  engine='openpyxl', mode='a') as writer:

 

            create_new_tables_sheet(writer, tables, sheetname)

 

    except Exception as e:

 

        print(f'[Myra input tables] : Failed to create table for {f}/n{e}')

 

 

 

# 2. 'myra' sheet

 

# We want to create a 96w based on Myra input.

 

# For each run ('star' files), remove the empties

 

def cleaning_run_empties(df, column_name):

 

    return df.replace('', None).dropna()[column_name].tolist()

 

 

 

# We want to create plates of 96w based on Myra input filled vertically.

 

#   -> Iterate through the runs and fill the plates. When we reach a plate end (H12)

 

#       we create a new one below the previous one etc..

 

pastel_colors = [

 

    "FAD02E", "F28D35", "D83367", "9A4F96", "2D3D6B", "3B96B1", "4CBF6F", "F5C6B8",

 

    "D9D8D1", "F1A7B5", "E9A6A6", "F2B0A1", "D4E6F1", "E2D1F9", "E7D0B0", "C6E2A2",

 

    "FFE156", "C9FF71", "E3F9DC", "B2D7E4"

 

]

 

 

 

def set_header_and_index(ws, start_row=0, end_row=10):

    for col in range(1, 13):  # Column headers 1-12

        ws.cell(row=start_row, column=col + 1, value=str(col))  

 

    idx = 0

    for row in range(start_row, end_row):  # Rows A-H (Index rows)

        ws.cell(row=row+1, column=1, value=chr(ord('A') + idx))

        idx = (idx + 1) if idx <= 6 else 0

 

def create_empty_table(ws, start_row=2, end_row=10):

    set_header_and_index(ws, start_row, end_row)

    blue_fill = PatternFill(start_color="87CEEB", fill_type="solid")

 

    for col in range(2, 14):  # Style header row (1-12)

 

        ws.cell(row=start_row, column=col).fill = blue_fill

 

 

 

    for row in range(start_row, end_row+1):  # Style index column (A-H)

 

        ws.cell(row=row, column=1).fill = blue_fill

 

 

 

def fill_cell(ws, row, col, val, color=None):

 

    if color:

 

        ws.cell(row=row, column=col, value=val).fill = color

 

    else:

 

        ws.cell(row=row, column=col, value=val)

 

 

 

def fill_well_plates(inp, wb):

 

    start_row = 3   # One row for the plate number, one for the header and we start from 3

 

    current_row, current_col = start_row, 1

 

    plate_number = 1

 

 

 

    ws = wb.create_sheet(title="Myra")  

 

   

 

    fill_cell(ws=ws, row=1, col=current_col, val=f"Plate {plate_number}")

 

   

 

    create_empty_table(ws)

 

   

 

    count = 0       # keep track of the row iteration

 

    for idx, (key, values) in enumerate(inp.items()):

 

        color = PatternFill(start_color=pastel_colors[idx], fill_type="solid")  

 

 

 

        for value in values:

 

 

 

            fill_cell(

 

                ws=ws,

 

                row=current_row,

 

                col=current_col + 1,

 

                val=value,

 

                color=color)

 

           

 

            current_row += 1

 

            count += 1

 

 

 

            if count == 8:  # If end of rows (A-H) is reached

                current_col += 1

                count = 0

            

                if (current_col > 12):

                    current_col = 1     # Start filling from the first column again

                    current_row += 1    # Add space for a new plate (space between tables)

                    plate_number += 1   # Increment plate number

 

                   

                    fill_cell(

                        ws=ws,

                        row=current_row,

                        col=current_col,

                        val=f"Plate {plate_number}")

 

                    current_row += 1    # Create the table below the plate number

                    

                    create_empty_table(

                        ws=ws,

                        start_row=current_row,

                        end_row=current_row + 8)

 

                    current_row += 1    # Starting to fill the table below the header

 

                    continue    # Go to the next iteration directly and don't reset

                                # the current row (to continue below the filled table another one)

 

                current_row = current_row - 8  # 'reset' the current row to fill anther column

 

def create_plates_myra(inp, filename):

    try:

        wb = load_workbook(filename)

        fill_well_plates(inp, wb)

        wb.save(filename)

 

    except Exception as e:

        logging.error(f"[create_plates_myra] Error occurred while creating the sheet: \n{e}")




##############################################################    

 

SERIAL_NUMBER_COL     = 'serial number'

SAMPLES_CODES_COL     = "samples codes"

PRODUCTION_DATE_COL   = "production date"

SOURCE_BARCODE_COL    = "source barcode" 

TARGET_POSITION_COL   = "target position"

SAMPLES_CODES_COL     = 'samples codes'

STAR_RUN_COL          = 'star run'

SOURCE_NAME_COL       = 'source name'

 

wd = os.path.dirname(os.path.realpath(__file__))

results_dir = os.path.join(wd, 'results')

logs_dir = os.path.join(wd, 'logs')

 

now = datetime.now()

formatted_date = now.strftime("%Y%m%d_%H%M%S")

 

create_dir(results_dir)

create_dir(logs_dir)

 

log_file = os.path.join(logs_dir, f'logs_{formatted_date}.txt')

logging.basicConfig(filename=log_file, level=logging.DEBUG, 

                    format='%(asctime)s - %(levelname)s - %(message)s')

 

data_dir        = os.path.join(wd, 'data')

results_file    = os.path.join(results_dir, f'results_{formatted_date}.xlsx')

filtered_file   = 'filtered_result'

concatened_file = 'concatenated_result'

 

# Get all run files (can be xls, csvs, or together)

runs_files = get_run_files(RUN_FILE_PATTERN, ('.xlsx', '.csv'))

 

if not runs_files:

    logging.error(f"[0] Error occurred while fetching 'star' files\n")

 

################################################################

 

### [0] - Setup ###

 

# If .csv, convert to xl (and remove csv files after convertion)

# We prefer to work with xl files

csv_files = [file for file in runs_files if file.endswith(".csv")]

 

if csv_files:

    for f in csv_files:

        try:

            convert_csv_to_xl(data_dir, f)

            os.remove(os.path.join(data_dir, f))

            logging.info(f"[1.3] Successfully converted {f} to excel.\n")

        except Exception as e:

            msg = f"[2.1] Error occurred while converting {f} to excel: {e}\n"

            logging.error(msg)

            exit(msg)

 

# After converting all files to xl, fetch again the run files with xl extention
runs_files = get_run_files(RUN_FILE_PATTERN, '.xlsx')

 
#################################################################
### [1]- Fetch 'Samples codes' and 'Production date' columns ###

try:
    for i in os.listdir(data_dir):
        if 'WGS' in i or'wgs' in i:
            sending_file = os.path.join(data_dir, i)
            break
    else:
        msg = 'Cannot find your sending file'
        logging.error(f"[1.1] {msg} \n")
        exit(msg) 
    logging.info(f"[1.1] Successfully fetched 'sending' file\n")
except Exception as e:
    logging.error(f"[1.1] Error occurred while fetching 'sending' file: {e}\n")

 

try:
    df = pd.read_excel(sending_file, header=None)

    # Reaching the table because it doesn't start from the first file row

    header_row = None

    for i, row in df.iterrows():
        row_str = row.astype(str).str.strip().str.lower()   # fix upper/lower case issue

        if SERIAL_NUMBER_COL in row_str.values:
            header_row = i
            break

    if header_row is not None:
        df = pd.read_excel(sending_file, header=header_row)
        df.columns = manage_columns_typo(df)

        null_rows = df[df[SAMPLES_CODES_COL].isna() | df[PRODUCTION_DATE_COL].isna()]

        if len(null_rows) > 0:
            msg = f'[1.2] Your data has null values:\n {null_rows}\n'
            logging.error(msg)
            exit(msg)

        columns = fetch_columns(df, [SAMPLES_CODES_COL, PRODUCTION_DATE_COL])

        df_filtered = df[columns]
        df_filtered = replace_hyphen_by_dot(df_filtered, SAMPLES_CODES_COL)
        
        logging.info(f"[1.2] Successfully fetching header form'sending' file\n")

    else:
        logging.error(f"[1.2] Error occurred while fetching header form 'sending' file: {e}\n")
        exit("Could not find the table header.")
except Exception as e:
    logging.error(f"[1.2] Error occurred while fetching header form'sending' file: {e}\n")
 

###########################################################
### [2] - Concate Star_0XX files with the sending file ###

df_results = df_filtered.copy()

try:
    for r in runs_files:
        df_results = fetch_and_concatenate(df_results, data_dir, r)
    logging.info(f"[2.1] Successfully fetching & concate 'star' files.\n")
except Exception as e:
    msg = f"[2.1] Error occurred while fetching & concate 'star' files: {e}\n"
    logging.error(msg)
    exit(msg)


df_results.columns = manage_columns_typo(df_results)

null_rows = df_results[df_results[SOURCE_BARCODE_COL].isna() | 
               df_results[TARGET_POSITION_COL].isna() |
               df_results[STAR_RUN_COL].isna()]

if len(null_rows) > 0:    
    msg = f'[2.2] Your data has null values: {null_rows}\n'
    logging.error(msg)
    exit(msg)

try:
    df_sorted = sort_df(df_results, STAR_RUN_COL)
    logging.info(f"[2.3] Successfully sorted values\n")

except Exception as e:
    msg = f"[2.3] Error occurred while sorting values: {e}\n"
    logging.error(msg)
    exit(msg)

 

try:

    with pd.ExcelWriter(results_file, engine='openpyxl') as writer:

        df_filtered.to_excel(writer, sheet_name=filtered_file, index=False)

        df_sorted.to_excel(writer, sheet_name=concatened_file, index=False)

        

        logging.info(f"[2.4] Successfully wrote excel table\n")

except Exception as e:
    msg = f"[2.4] Error occurred while writing excel table: {e}\n"
    logging.error(msg)
    exit(msg)

##############################################################
### [3] - Create Tables 96w like, and highlighted ###
# create like 96w table A1, B1, C1 ....
# color the user's samples

try:
    create_tables(data_dir, df_sorted, results_file)
    logging.info('[2.5] Successfully created tables sheets (plates, long table)\n')
    msg = 'The script ran successfully\n\n--- THE END ---'
    logging.info(msg)
    print(msg)

except Exception as e:
    logging.error(f"[2.5] Error occurred while creating tables sheets: {e}")
